#!/usr/bin/env python3
"""AI介绍 批量重写管线。

对 `.plbackup`（SwiftData 的 SQLite store）里的 `ZBOOK.ZBOOKINTRODUCTION` 做
「存档 → 清空 → 分批 → 合并校验 → 写回 → checkpoint → 导出」的一条流水线。

绝不改动原始备份文件：所有操作都在 work 副本上做，original 副本只用于最终比对。
用法见 .claude/skills/ai-book-intro/SKILL.md。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sqlite3
import sys
from dataclasses import dataclass
from pathlib import Path

# Core Data 的时间戳纪元是 2001-01-01，Unix 纪元是 1970-01-01
CORE_DATA_EPOCH_OFFSET = 978_307_200

# 豆瓣简介作为原料的截断上限：写一段 900 字的介绍用不到更多，
# 而库里最长的有 7158 字，全塞进批次文件会让 agent 输入无谓膨胀。
DOUBAN_INTRO_CAP = 2500

# 分轨门槛：豆瓣简介少于这个字数视为「缺料」，需要联网查证
TRACK_A_MIN_DOUBAN = 50

BATCH_SIZE_TRACK_A = 25
BATCH_SIZE_TRACK_B = 10

WORKDIR = Path("/tmp/pl-ai-intro")


# --------------------------------------------------------------------------- #
# 基础设施
# --------------------------------------------------------------------------- #


def connect(db: Path, *, readonly: bool = False) -> sqlite3.Connection:
    uri = f"file:{db}?mode=ro" if readonly else f"file:{db}"
    conn = sqlite3.connect(uri, uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def publish_year(raw: float | None) -> int | None:
    """Core Data TIMESTAMP → 公历年份。

    库里有脏数据（实测有一条换算出公元 20208 年），超出合理区间的一律当作没有年份 ——
    宁可导语段不写年份，也不能把假年份喂给 agent 当事实。
    """
    if raw is None:
        return None
    import datetime

    try:
        ts = float(raw) + CORE_DATA_EPOCH_OFFSET
        year = datetime.datetime.fromtimestamp(ts, datetime.timezone.utc).year
    except (ValueError, OverflowError, OSError):
        return None
    return year if 1000 <= year <= datetime.date.today().year + 1 else None


def batch_path(kind: str, batch_id: str) -> Path:
    return WORKDIR / kind / f"batch_{batch_id}.json"


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


# --------------------------------------------------------------------------- #
# 1. archive —— 把现有 AI介绍 存档，迁移出错时的唯一退路
# --------------------------------------------------------------------------- #


def cmd_archive(args: argparse.Namespace) -> int:
    conn = connect(args.db, readonly=True)
    rows = conn.execute(
        """
        SELECT Z_PK, ZTITLE, ZAUTHOR, ZPUBLISHER, ZISBN, ZWEREADBOOKID, ZBOOKINTRODUCTION
        FROM ZBOOK
        WHERE ZBOOKINTRODUCTION IS NOT NULL AND ZBOOKINTRODUCTION <> ''
        ORDER BY Z_PK
        """
    ).fetchall()
    conn.close()

    records = [
        {
            "pk": r["Z_PK"],
            "title": r["ZTITLE"],
            "author": r["ZAUTHOR"],
            "publisher": r["ZPUBLISHER"],
            "isbn": r["ZISBN"],
            "weread_id": r["ZWEREADBOOKID"],
            "intro": r["ZBOOKINTRODUCTION"],
        }
        for r in rows
    ]
    write_json(args.out_json, records)
    _write_intro_xlsx(args.out_xlsx, records, sheet_title="旧AI介绍存档")
    print(f"存档 {len(records)} 条 → {args.out_json}")
    print(f"存档 {len(records)} 条 → {args.out_xlsx}")
    return 0


# --------------------------------------------------------------------------- #
# 2. clear —— 清空字段，并验证真的清空了
# --------------------------------------------------------------------------- #


def cmd_clear(args: argparse.Namespace) -> int:
    conn = connect(args.db)
    before = conn.execute(
        "SELECT COUNT(*) FROM ZBOOK WHERE ZBOOKINTRODUCTION IS NOT NULL AND ZBOOKINTRODUCTION <> ''"
    ).fetchone()[0]
    conn.execute("UPDATE ZBOOK SET ZBOOKINTRODUCTION = NULL")
    conn.commit()
    after = conn.execute(
        "SELECT COUNT(*) FROM ZBOOK WHERE ZBOOKINTRODUCTION IS NOT NULL AND ZBOOKINTRODUCTION <> ''"
    ).fetchone()[0]
    conn.close()

    print(f"清空前 {before} 条，清空后 {after} 条")
    if after != 0:
        print("清空未生效，中止", file=sys.stderr)
        return 1
    return 0


# --------------------------------------------------------------------------- #
# 3. export-batches —— 按轨道切批，产出 agent 的输入文件
# --------------------------------------------------------------------------- #


@dataclass
class BookRow:
    pk: int
    title: str
    author: str
    publisher: str | None
    year: int | None
    isbn: str | None
    book_type: str | None
    douban_intro: str
    author_intro: str

    @property
    def track(self) -> str:
        return "A" if len(self.douban_intro) >= TRACK_A_MIN_DOUBAN else "B"

    def to_payload(self) -> dict:
        intro = self.douban_intro
        truncated = len(intro) > DOUBAN_INTRO_CAP
        payload = {
            "pk": self.pk,
            "title": self.title,
            "author": self.author,
            "publisher": self.publisher or None,
            "year": self.year,
            "isbn": self.isbn or None,
            "book_type": self.book_type or None,
            "douban_intro": intro[:DOUBAN_INTRO_CAP],
            "author_intro": self.author_intro[:DOUBAN_INTRO_CAP],
        }
        if truncated:
            payload["douban_intro_truncated"] = True
        return payload


def load_books(db: Path, *, only_empty: bool) -> list[BookRow]:
    conn = connect(db, readonly=True)
    where = ""
    if only_empty:
        where = "WHERE ZBOOKINTRODUCTION IS NULL OR ZBOOKINTRODUCTION = ''"
    rows = conn.execute(
        f"""
        SELECT Z_PK, ZTITLE, ZAUTHOR, ZPUBLISHER, ZPUBLISHDATE, ZISBN, ZBOOKTYPE,
               ZBOOKDESCRIPTION, ZAUTHORDESCRIPTION
        FROM ZBOOK {where} ORDER BY Z_PK
        """
    ).fetchall()
    conn.close()
    return [
        BookRow(
            pk=r["Z_PK"],
            title=r["ZTITLE"] or "",
            author=r["ZAUTHOR"] or "",
            publisher=r["ZPUBLISHER"],
            year=publish_year(r["ZPUBLISHDATE"]),
            isbn=r["ZISBN"],
            book_type=r["ZBOOKTYPE"],
            douban_intro=(r["ZBOOKDESCRIPTION"] or "").strip(),
            author_intro=(r["ZAUTHORDESCRIPTION"] or "").strip(),
        )
        for r in rows
    ]


def chunk(items: list, size: int) -> list[list]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def cmd_export_batches(args: argparse.Namespace) -> int:
    books = load_books(args.db, only_empty=args.only_empty)
    track_a = [b for b in books if b.track == "A"]
    track_b = [b for b in books if b.track == "B"]

    batches: list[tuple[str, list[BookRow]]] = []
    for group in chunk(track_a, BATCH_SIZE_TRACK_A):
        batches.append(("A", group))
    for group in chunk(track_b, BATCH_SIZE_TRACK_B):
        batches.append(("B", group))

    manifest = []
    for index, (track, group) in enumerate(batches, start=1):
        batch_id = f"{index:03d}"
        write_json(
            batch_path("in", batch_id),
            {
                "batch_id": batch_id,
                "track": track,
                "books": [b.to_payload() for b in group],
            },
        )
        manifest.append({"batch_id": batch_id, "track": track, "count": len(group)})

    write_json(WORKDIR / "manifest.json", manifest)
    print(f"共 {len(books)} 本：track A {len(track_a)} 本、track B {len(track_b)} 本")
    print(f"切成 {len(batches)} 批 → {WORKDIR / 'in'}")
    print(f"清单 → {WORKDIR / 'manifest.json'}")
    return 0


# --------------------------------------------------------------------------- #
# 4. merge —— 收集 agent 输出、逐条校验、产出待写回集合与失败清单
# --------------------------------------------------------------------------- #


def cmd_merge(args: argparse.Namespace) -> int:
    sys.path.insert(0, str(Path(__file__).parent))
    # 延迟导入，让前几步不依赖它
    from validate import duplicate_pks, normalize_indent, validate_intro

    manifest = json.loads((WORKDIR / "manifest.json").read_text(encoding="utf-8"))
    inputs = {
        m["batch_id"]: json.loads(
            batch_path("in", m["batch_id"]).read_text(encoding="utf-8")
        )
        for m in manifest
    }

    accepted: list[dict] = []
    insufficient: list[dict] = []
    rejected: list[dict] = []
    missing_batches: list[str] = []

    for m in manifest:
        batch_id = m["batch_id"]
        out_path = batch_path("out", batch_id)
        if not out_path.exists():
            missing_batches.append(batch_id)
            continue
        try:
            payload = json.loads(out_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            rejected.append(
                {"batch_id": batch_id, "pk": None, "reasons": [f"JSON 解析失败: {exc}"]}
            )
            missing_batches.append(batch_id)
            continue

        by_pk = {b["pk"]: b for b in inputs[batch_id]["books"]}
        results_in_batch = payload.get("results", [])
        for dupe_pk in duplicate_pks(
            r.get("pk") for r in results_in_batch if r.get("pk") is not None
        ):
            rejected.append(
                {
                    "batch_id": batch_id,
                    "pk": dupe_pk,
                    "title": by_pk.get(dupe_pk, {}).get("title", ""),
                    "reasons": ["同一个 pk 写了多段（并发撞车），需删到只剩一段"],
                }
            )
        seen: set[int] = set()
        for result in results_in_batch:
            pk = result.get("pk")
            source = by_pk.get(pk)
            if source is None:
                rejected.append(
                    {"batch_id": batch_id, "pk": pk, "reasons": ["pk 不在该批次输入里"]}
                )
                continue
            seen.add(pk)
            if result.get("status") == "insufficient_data":
                insufficient.append(
                    {
                        "batch_id": batch_id,
                        "pk": pk,
                        "title": source["title"],
                        "note": result.get("note", ""),
                    }
                )
                continue

            reasons = validate_intro(
                result.get("intro", ""),
                title=source["title"],
                douban_intro=source.get("douban_intro", ""),
            )
            if reasons:
                rejected.append(
                    {
                        "batch_id": batch_id,
                        "pk": pk,
                        "title": source["title"],
                        "reasons": reasons,
                    }
                )
            else:
                # 缩进统一成标杆样板的半角空格：agent 可能用全角/多空格/tab，
                # 2904 条里混着几种缩进在详情页上很难看
                accepted.append(
                    {
                        "pk": pk,
                        "title": source["title"],
                        "intro": normalize_indent(result["intro"]),
                    }
                )

        for pk in sorted(set(by_pk) - seen):
            rejected.append(
                {
                    "batch_id": batch_id,
                    "pk": pk,
                    "title": by_pk[pk]["title"],
                    "reasons": ["agent 输出里漏了这本"],
                }
            )

    write_json(WORKDIR / "merged.json", accepted)
    write_json(
        WORKDIR / "merge_report.json",
        {
            "accepted": len(accepted),
            "insufficient": insufficient,
            "rejected": rejected,
            "missing_batches": missing_batches,
        },
    )

    print(f"通过 {len(accepted)} 条")
    print(f"资料不足留空 {len(insufficient)} 条")
    print(f"打回 {len(rejected)} 条")
    if missing_batches:
        print(f"缺输出的批次 {len(missing_batches)} 个: {', '.join(missing_batches)}")
    print(f"→ {WORKDIR / 'merged.json'} / {WORKDIR / 'merge_report.json'}")
    return 0


# --------------------------------------------------------------------------- #
# 3b. status —— 一眼看全局进度（跑上百批时反复要看）
# --------------------------------------------------------------------------- #


def cmd_status(args: argparse.Namespace) -> int:
    sys.path.insert(0, str(Path(__file__).parent))
    from validate import (
        PLAGIARISM_RUN,
        char_count,
        longest_common_run,
        strip_for_compare,
        title_mismatch,
        validate_intro,
    )

    # 「贴线」审计区间：红线内但仍可能是照搬。实测有 agent 靠删一个字
    # （`所能产生`→`所产生`、插入一个`都`）把 40+ 压到 30+ 蒙过红线。
    # 但 30–39 也确实有不可回避的合理残留 —— 机构专名并列、书名加版次、
    # 带引号的具名引文。所以只列出来供抽查，不打回。
    near_floor = 30

    manifest = json.loads((WORKDIR / "manifest.json").read_text(encoding="utf-8"))

    done_batches = 0
    total_books = sum(m["count"] for m in manifest)
    drafted = validated = rejected = 0
    lengths: list[int] = []
    problems: list[str] = []
    mismatches: list[str] = []
    near_line: list[str] = []
    # 「资料不足」留空是契约里的合法结果（ingest-text / merge 都这么认）。
    # 早先 status 对空正文跑 validate_intro，把它们混进「待修（硬打回）」，
    # 同一份稿件三个命令给出三种结论，跑批时会误判进度。
    insufficient_rows: list[str] = []

    for m in manifest:
        batch_id, expected = m["batch_id"], m["count"]
        draft = WORKDIR / "text" / f"batch_{batch_id}.txt"
        out = batch_path("out", batch_id)

        n_draft = 0
        if draft.exists():
            n_draft = len(parse_text_draft(draft.read_text(encoding="utf-8")))
        drafted += n_draft

        if not out.exists():
            if n_draft and args.verbose:
                print(
                    f"  batch_{batch_id} ({m['track']}): 稿件 {n_draft}/{expected}，未 ingest"
                )
            continue

        src = {
            b["pk"]: b
            for b in json.loads(batch_path("in", batch_id).read_text(encoding="utf-8"))[
                "books"
            ]
        }
        results = json.loads(out.read_text(encoding="utf-8"))["results"]
        bad = 0
        skipped = 0
        for r in results:
            book = src.get(r["pk"], {})
            if r.get("status") == "insufficient_data":
                skipped += 1
                insufficient_rows.append(
                    f"  batch_{batch_id} pk={r['pk']} {book.get('title', '?')[:16]}"
                )
                continue
            reasons = validate_intro(
                r.get("intro", ""),
                title=book.get("title", ""),
                douban_intro=book.get("douban_intro", ""),
            )
            lengths.append(char_count(r.get("intro", "")))
            if reasons:
                bad += 1
                problems.append(
                    f"  batch_{batch_id} pk={r['pk']} {book.get('title', '?')[:16]}: "
                    f"{'; '.join(reasons)[:100]}"
                )
            note = title_mismatch(
                book.get("title", ""), r.get("intro", "").split("\n")[0]
            )
            if note:
                mismatches.append(f"  batch_{batch_id} pk={r['pk']}: {note}")

            douban = book.get("douban_intro", "")
            if douban and r.get("intro"):
                run = longest_common_run(
                    strip_for_compare(r["intro"]), strip_for_compare(douban)
                )
                if near_floor <= run < PLAGIARISM_RUN:
                    near_line.append(
                        f"  batch_{batch_id} pk={r['pk']} "
                        f"{book.get('title', '?')[:16]}: 重合 {run} 字"
                    )
        validated += len(results) - bad - skipped
        rejected += bad
        if len(results) == expected and bad == 0:
            done_batches += 1
        elif args.verbose:
            print(
                f"  batch_{batch_id} ({m['track']}): {len(results)}/{expected} 段，打回 {bad}"
            )

    print(f"批次   {done_batches}/{len(manifest)} 批完成且全通过")
    print(
        f"藏书   {validated}/{total_books} 本已通过校验（打回 {rejected}，稿件累计 {drafted} 段）"
    )
    if lengths:
        print(
            f"字数   最短 {min(lengths)}，最长 {max(lengths)}，"
            f"平均 {sum(lengths) // len(lengths)}（标杆样板 694）"
        )
    if insufficient_rows:
        print(
            f"\n资料不足留空 {len(insufficient_rows)} 条（契约认可的合法结果，不是待修）:"
        )
        for line in insufficient_rows[: args.limit]:
            print(line)
        if len(insufficient_rows) > args.limit:
            print(f"  …另有 {len(insufficient_rows) - args.limit} 条")
    if problems:
        print(f"\n待修 {len(problems)} 条（硬打回）:")
        for line in problems[: args.limit]:
            print(line)
        if len(problems) > args.limit:
            print(f"  …另有 {len(problems) - args.limit} 条，加 --limit 看更多")
    if near_line:
        print(
            f"\n贴线待抽查 {len(near_line)} 条"
            f"（重合 {near_floor}–{PLAGIARISM_RUN - 1} 字，不打回）:"
        )
        for line in near_line[: args.limit]:
            print(line)
        if len(near_line) > args.limit:
            print(f"  …另有 {len(near_line) - args.limit} 条，加 --limit 看更多")
    if mismatches:
        print(
            f"\n待人工确认 {len(mismatches)} 条（书名对不上，多为繁简/中译名，不打回）:"
        )
        for line in mismatches[: args.limit]:
            print(line)
        if len(mismatches) > args.limit:
            print(f"  …另有 {len(mismatches) - args.limit} 条，加 --limit 看更多")
    return 0


# --------------------------------------------------------------------------- #
# 4b. ingest-text —— 把纯文本稿转成 out/batch_NNN.json
# --------------------------------------------------------------------------- #

# 纯文本稿的分隔行：`### <pk>`
PK_HEADER = re.compile(r"^###\s+(\d+)\s*$")


def parse_text_draft(text: str) -> list[dict]:
    """解析 `### <pk>` 分隔的纯文本稿。

    直接手写 JSON 要为每段正文转义几十个 \\n，既费 token 又容易写坏；
    纯文本稿由脚本转 JSON，转义交给 json.dumps。
    """
    entries: list[dict] = []
    pk: int | None = None
    buffer: list[str] = []

    def flush() -> None:
        if pk is None:
            return
        body = "\n".join(buffer).strip("\n")
        if body:
            entries.append({"pk": pk, "intro": body})

    for line in text.split("\n"):
        match = PK_HEADER.match(line)
        if match:
            flush()
            pk = int(match.group(1))
            buffer = []
        elif pk is not None:
            buffer.append(line)
    flush()
    return entries


def cmd_ingest_text(args: argparse.Namespace) -> int:
    """转 JSON **并当场校验**。

    agent 的自检循环就跑这一条命令，所以质量闸门必须长在这里 —— 实测：
    自发跑过校验器的两个批次 0 打回，没跑的三个批次出了 9 条抄袭。
    只报"缺几段"不报"哪段不合格"，等于把问题推到事后返工。
    """
    sys.path.insert(0, str(Path(__file__).parent))
    from validate import duplicate_pks, title_mismatch, validate_intro

    batch_id = args.batch_id
    draft = args.draft or (WORKDIR / "text" / f"batch_{batch_id}.txt")
    source = json.loads(batch_path("in", batch_id).read_text(encoding="utf-8"))
    books = {b["pk"]: b for b in source["books"]}

    entries = parse_text_draft(draft.read_text(encoding="utf-8"))
    seen = {e["pk"] for e in entries}

    from validate import is_insufficient

    results = [
        {
            "pk": e["pk"],
            "title": books.get(e["pk"], {}).get("title", ""),
            "status": "insufficient_data" if is_insufficient(e["intro"]) else "ok",
            "intro": "" if is_insufficient(e["intro"]) else e["intro"],
            **({"note": e["intro"].strip()} if is_insufficient(e["intro"]) else {}),
        }
        for e in entries
    ]
    write_json(batch_path("out", batch_id), {"batch_id": batch_id, "results": results})

    missing = sorted(set(books) - seen)
    extra = sorted(seen - set(books))
    dupes = duplicate_pks(e["pk"] for e in entries)
    print(f"批次 {batch_id}: 输入 {len(books)} 本，稿件 {len(entries)} 段")
    if missing:
        print(f"缺 {len(missing)} 段: {missing}")
    if extra:
        print(f"多出（pk 不在该批次）: {extra}")
    if dupes:
        print(f"重复 {len(dupes)} 个 pk（同一本写了多段，要删到只剩一段）: {dupes}")
        print("  多半是两个 agent 并发写了同一份稿件。merge 会让后写的那份静默覆盖，")
        print("  所以必须在这一步修掉，别指望下游发现。")

    problems: list[str] = []
    insufficient = 0
    for r in results:
        if r["status"] == "insufficient_data":
            insufficient += 1
            continue
        book = books.get(r["pk"], {})
        for reason in validate_intro(
            r["intro"],
            title=book.get("title", ""),
            douban_intro=book.get("douban_intro", ""),
        ):
            problems.append(f"  pk={r['pk']} {book.get('title', '?')[:20]}: {reason}")

    if insufficient:
        print(f"其中 {insufficient} 段标为「资料不足」留空（track B 的合法结果）")

    if problems:
        print(f"\n不合格 {len(problems)} 条，必须改掉:")
        for line in problems:
            print(line)
        print("\n提示：「与豆瓣简介连续重合 N 字」= 抄了原料里的句子，把那一条重写；")
        print("     「篇幅过短」= 把该段补到 700 字以上。改完再跑本命令。")
    else:
        print("校验：全部合格")

    for r in results:
        note = title_mismatch(
            books.get(r["pk"], {}).get("title", ""), r["intro"].split("\n")[0]
        )
        if note:
            print(f"（供参考，不必改）pk={r['pk']}: {note}")

    return 0 if not missing and not extra and not problems and not dupes else 1


# --------------------------------------------------------------------------- #
# 5. write-db —— 按主键精确写回
# --------------------------------------------------------------------------- #


def cmd_write_db(args: argparse.Namespace) -> int:
    records = json.loads(args.merged.read_text(encoding="utf-8"))
    conn = connect(args.db)
    updated = 0
    for record in records:
        cursor = conn.execute(
            "UPDATE ZBOOK SET ZBOOKINTRODUCTION = ? WHERE Z_PK = ?",
            (record["intro"], record["pk"]),
        )
        updated += cursor.rowcount
    conn.commit()
    filled = conn.execute(
        "SELECT COUNT(*) FROM ZBOOK WHERE ZBOOKINTRODUCTION IS NOT NULL AND ZBOOKINTRODUCTION <> ''"
    ).fetchone()[0]
    conn.close()
    print(f"写回 {updated} 行；库内非空 AI介绍 现为 {filled} 条")
    return 0 if updated == len(records) else 1


# --------------------------------------------------------------------------- #
# 6. checkpoint —— 把 WAL 并回主文件，保证成品是自包含单文件
# --------------------------------------------------------------------------- #


def cmd_checkpoint(args: argparse.Namespace) -> int:
    conn = connect(args.db)
    mode = conn.execute("PRAGMA journal_mode").fetchone()[0]
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    conn.commit()
    conn.close()

    removed = []
    for suffix in ("-wal", "-shm"):
        sidecar = Path(str(args.db) + suffix)
        if sidecar.exists():
            size = sidecar.stat().st_size
            if size > 0 and suffix == "-wal":
                print(f"{sidecar.name} 仍有 {size} 字节未并入，中止", file=sys.stderr)
                return 1
            sidecar.unlink()
            removed.append(sidecar.name)

    print(f"journal_mode={mode}，checkpoint 完成，清掉边车: {removed or '无'}")
    return 0


# --------------------------------------------------------------------------- #
# 7. export-xlsx —— 给用户下载的成品表
# --------------------------------------------------------------------------- #


def _write_intro_xlsx(path: Path, records: list[dict], *, sheet_title: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    # 前 5 个表头必须与 ExcelImportExportService.parseIntroductionEntries 读的列名完全一致，
    # 否则 App 的「导入 AI介绍」认不出来
    headers = ["书名", "作者", "出版社", "ISBN", "微信读书ID", "AI介绍", "字数"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for record in records:
        intro = record.get("intro") or ""
        ws.append(
            [
                record.get("title") or "",
                record.get("author") or "",
                record.get("publisher") or "",
                record.get("isbn") or "",
                record.get("weread_id") or "",
                intro,
                len(intro),
            ]
        )

    widths = {"A": 32, "B": 20, "C": 22, "D": 18, "E": 14, "F": 100, "G": 8}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def cmd_export_xlsx(args: argparse.Namespace) -> int:
    conn = connect(args.db, readonly=True)
    rows = conn.execute(
        """
        SELECT ZTITLE, ZAUTHOR, ZPUBLISHER, ZISBN, ZWEREADBOOKID, ZBOOKINTRODUCTION
        FROM ZBOOK ORDER BY Z_PK
        """
    ).fetchall()
    conn.close()

    records = [
        {
            "title": r["ZTITLE"],
            "author": r["ZAUTHOR"],
            "publisher": r["ZPUBLISHER"],
            "isbn": r["ZISBN"],
            "weread_id": r["ZWEREADBOOKID"],
            "intro": r["ZBOOKINTRODUCTION"] or "",
        }
        for r in rows
    ]
    if args.skip_empty:
        records = [r for r in records if r["intro"]]

    _write_intro_xlsx(args.out, records, sheet_title="AI介绍")
    print(f"导出 {len(records)} 行 → {args.out}")
    return 0


# --------------------------------------------------------------------------- #
# 8. verify —— 证明只有 AI介绍 变了，别的字段一个都没动
# --------------------------------------------------------------------------- #


UNTOUCHED_COLUMNS = [
    "ZTITLE",
    "ZAUTHOR",
    "ZBOOKDESCRIPTION",
    "ZAUTHORDESCRIPTION",
    "ZISBN",
    "ZPUBLISHER",
    "ZNOTES",
    "ZSTATUS",
    "ZISARCHIVED",
    "ZRATING",
    "ZCURRENTPAGE",
    "ZTOTALPAGES",
    "ZWEREADBOOKID",
    "ZCOVERIMAGEDATA",
]


def cmd_verify(args: argparse.Namespace) -> int:
    # 边车检查必须在**打开数据库之前**做：库是 WAL 模式，
    # 连上去（哪怕只读）就会重新生成 -wal/-shm，放在后面查等于自己制造失败。
    stale_sidecars = [
        Path(str(args.db) + suffix).name
        for suffix in ("-wal", "-shm")
        if Path(str(args.db) + suffix).exists()
        and Path(str(args.db) + suffix).stat().st_size > 0
    ]

    conn = connect(args.db, readonly=True)
    conn.execute("ATTACH DATABASE ? AS orig", (f"file:{args.original}?mode=ro",))

    failures: list[str] = []

    counts = conn.execute(
        "SELECT (SELECT COUNT(*) FROM ZBOOK), (SELECT COUNT(*) FROM orig.ZBOOK)"
    ).fetchone()
    if counts[0] != counts[1]:
        failures.append(f"行数不一致: 成品 {counts[0]} vs 原始 {counts[1]}")
    print(f"行数: 成品 {counts[0]}，原始 {counts[1]}")

    for column in UNTOUCHED_COLUMNS:
        diff = conn.execute(
            f"""
            SELECT COUNT(*) FROM ZBOOK w JOIN orig.ZBOOK o ON w.Z_PK = o.Z_PK
            WHERE w.{column} IS NOT o.{column}
            """
        ).fetchone()[0]
        if diff:
            failures.append(f"{column} 有 {diff} 行被改动")
        print(f"{column:24s} 差异 {diff} 行")

    changed = conn.execute(
        """
        SELECT COUNT(*) FROM ZBOOK w JOIN orig.ZBOOK o ON w.Z_PK = o.Z_PK
        WHERE w.ZBOOKINTRODUCTION IS NOT o.ZBOOKINTRODUCTION
        """
    ).fetchone()[0]
    filled = conn.execute(
        "SELECT COUNT(*) FROM ZBOOK WHERE ZBOOKINTRODUCTION IS NOT NULL AND ZBOOKINTRODUCTION <> ''"
    ).fetchone()[0]
    stats = conn.execute(
        """
        SELECT MIN(LENGTH(ZBOOKINTRODUCTION)), MAX(LENGTH(ZBOOKINTRODUCTION)),
               CAST(AVG(LENGTH(ZBOOKINTRODUCTION)) AS INT)
        FROM ZBOOK WHERE ZBOOKINTRODUCTION IS NOT NULL AND ZBOOKINTRODUCTION <> ''
        """
    ).fetchone()
    conn.close()

    print(f"ZBOOKINTRODUCTION 变更 {changed} 行；非空 {filled} 条")
    print(f"新 AI介绍 字数: 最短 {stats[0]}，最长 {stats[1]}，平均 {stats[2]}")

    for name in stale_sidecars:
        failures.append(f"成品旁残留非空边车 {name}，恢复时会丢改动（先跑 checkpoint）")

    if failures:
        print("\n校验失败:", file=sys.stderr)
        for failure in failures:
            print(f"  - {failure}", file=sys.stderr)
        return 1
    print("\n校验通过：只有 AI介绍 变了，无边车残留")
    return 0


# --------------------------------------------------------------------------- #
# 9. publish —— 复制成品到 iCloud 目录
# --------------------------------------------------------------------------- #


def cmd_publish(args: argparse.Namespace) -> int:
    args.dest.mkdir(parents=True, exist_ok=True)
    for source in args.files:
        target = args.dest / source.name
        shutil.copy2(source, target)
        size = os.path.getsize(target) / 1_048_576
        print(f"{target}  ({size:.1f} MB)")
    return 0


# --------------------------------------------------------------------------- #


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)

    work = WORKDIR / "work.sqlite"

    p = sub.add_parser("archive", help="导出现有 AI介绍 存档")
    p.add_argument("--db", type=Path, default=work)
    p.add_argument(
        "--out-json", type=Path, default=WORKDIR / "archive" / "old_intros.json"
    )
    p.add_argument(
        "--out-xlsx", type=Path, default=WORKDIR / "archive" / "旧AI介绍_存档.xlsx"
    )
    p.set_defaults(func=cmd_archive)

    p = sub.add_parser("clear", help="清空 ZBOOKINTRODUCTION")
    p.add_argument("--db", type=Path, default=work)
    p.set_defaults(func=cmd_clear)

    p = sub.add_parser("export-batches", help="切批产出 agent 输入")
    p.add_argument("--db", type=Path, default=work)
    p.add_argument(
        "--only-empty",
        action="store_true",
        help="只处理 AI介绍 为空的书（增量补新书用）",
    )
    p.set_defaults(func=cmd_export_batches)

    p = sub.add_parser("status", help="一眼看全局进度与待修条目")
    p.add_argument(
        "-v", "--verbose", action="store_true", help="逐批列出未完成/有打回的批次"
    )
    p.add_argument(
        "--limit", type=int, default=15, help="最多列出多少条待修（默认 15）"
    )
    p.set_defaults(func=cmd_status)

    p = sub.add_parser(
        "ingest-text", help="把 `### <pk>` 分隔的纯文本稿转成 out/batch_NNN.json"
    )
    p.add_argument("batch_id")
    p.add_argument("--draft", type=Path, default=None)
    p.set_defaults(func=cmd_ingest_text)

    p = sub.add_parser("merge", help="收集校验 agent 输出")
    p.set_defaults(func=cmd_merge)

    p = sub.add_parser("write-db", help="按主键写回数据库")
    p.add_argument("--db", type=Path, default=work)
    p.add_argument("--merged", type=Path, default=WORKDIR / "merged.json")
    p.set_defaults(func=cmd_write_db)

    p = sub.add_parser("checkpoint", help="WAL 并回主文件并清边车")
    p.add_argument("--db", type=Path, default=work)
    p.set_defaults(func=cmd_checkpoint)

    p = sub.add_parser("export-xlsx", help="导出成品 xlsx")
    p.add_argument("--db", type=Path, default=work)
    p.add_argument("--out", type=Path, default=WORKDIR / "AI介绍.xlsx")
    p.add_argument("--skip-empty", action="store_true", help="不导出 AI介绍 为空的行")
    p.set_defaults(func=cmd_export_xlsx)

    p = sub.add_parser("verify", help="与原始库比对，证明只有 AI介绍 变了")
    p.add_argument("--db", type=Path, default=work)
    p.add_argument("--original", type=Path, default=WORKDIR / "original.sqlite")
    p.set_defaults(func=cmd_verify)

    p = sub.add_parser("publish", help="复制成品到 iCloud 目录")
    p.add_argument("files", type=Path, nargs="+")
    p.add_argument("--dest", type=Path, required=True)
    p.set_defaults(func=cmd_publish)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
